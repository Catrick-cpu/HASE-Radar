"""
AERIS-10P Spreadsheet Generator
=================================
Generates formatted Excel .xlsx files from the CSV data in this directory.

Requirements:
    pip install openpyxl

Usage:
    python generate_spreadsheets.py

Outputs:
    BOM.xlsx      — Bill of Materials with formatted table
    Budget.xlsx   — Budget breakdown with category totals
    Timeline.xlsx — Gantt-style project timeline
"""

import csv
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                   numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not found. Install with: pip install openpyxl")
    print("Continuing with CSV output only...")


SCRIPT_DIR = Path(__file__).parent

# ─── Style helpers ────────────────────────────────────────────────────────────

def header_style():
    return {
        'font':      Font(bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill("solid", fgColor="1F497D"),
        'alignment': Alignment(horizontal='center', wrap_text=True),
    }

def apply_style(cell, style_dict):
    for attr, val in style_dict.items():
        setattr(cell, attr, val)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def alt_fill(row_num):
    color = "EBF0F8" if row_num % 2 == 0 else "FFFFFF"
    return PatternFill("solid", fgColor=color)

def money_format(cell):
    cell.number_format = '"$"#,##0.00'

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

# ─── BOM.xlsx ─────────────────────────────────────────────────────────────────

def create_bom_xlsx():
    if not HAS_OPENPYXL:
        return
    print("Creating BOM.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.freeze_panes = "A2"

    # Read CSV
    csv_path = SCRIPT_DIR / "BOM.csv"
    if not csv_path.exists():
        print(f"  Warning: {csv_path} not found, skipping BOM.xlsx")
        return

    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return

    headers = rows[0]
    hdr = header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        apply_style(cell, hdr)

    # Data rows
    total_cost = 0.0
    for ri, row in enumerate(rows[1:], 2):
        fill = alt_fill(ri)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            # Format price columns
            if headers[ci-1] in ('Unit_Price_USD', 'Total_Price_USD'):
                try:
                    cell.value = float(val)
                    money_format(cell)
                    if headers[ci-1] == 'Total_Price_USD':
                        total_cost += float(val)
                except (ValueError, TypeError):
                    pass

    # Total row
    total_row = len(rows) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, h in enumerate(headers, 1):
        if h == 'Total_Price_USD':
            cell = ws.cell(row=total_row, column=ci, value=total_cost)
            cell.font = Font(bold=True)
            money_format(cell)
            cell.fill = PatternFill("solid", fgColor="C6EFCE")

    # Column widths
    widths = {'Line': 6, 'Category': 14, 'Ref_Des': 12, 'Description': 40,
              'Manufacturer': 20, 'MPN': 22, 'Supplier': 12,
              'Supplier_PN': 18, 'Qty': 6, 'Unit_Price_USD': 14,
              'Total_Price_USD': 14, 'Package': 14, 'Notes': 35, 'Alternative_MPN': 18}
    for ci, h in enumerate(headers, 1):
        set_col_width(ws, ci, widths.get(h, 12))

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "AERIS-10P BOM Summary"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'] = "Total Component Cost"
    ws2['B3'] = total_cost
    money_format(ws2['B3'])
    ws2['B3'].font = Font(bold=True)
    ws2['A4'] = "Budget Limit"
    ws2['B4'] = 8000.00
    money_format(ws2['B4'])
    ws2['A5'] = "Remaining Budget"
    ws2['B5'] = 8000.0 - total_cost
    money_format(ws2['B5'])
    ws2['B5'].font = Font(bold=True, color="006100")

    wb.save(SCRIPT_DIR / "BOM.xlsx")
    print(f"  Saved BOM.xlsx ({len(rows)-1} items, total ${total_cost:.2f})")


# ─── Budget.xlsx ──────────────────────────────────────────────────────────────

def create_budget_xlsx():
    if not HAS_OPENPYXL:
        return
    print("Creating Budget.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.freeze_panes = "A2"

    csv_path = SCRIPT_DIR / "Budget.csv"
    if not csv_path.exists():
        print(f"  Warning: {csv_path} not found")
        return

    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))

    headers = rows[0]
    hdr = header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        apply_style(cell, hdr)

    category_colors = {
        'Electronics': "FFF2CC",
        'PCBs':        "DDEBF7",
        'Cables':      "E2EFDA",
        'Mechanical':  "FCE4D6",
        'Test Equipment': "EAD1DC",
        'Consumables': "F4CCFF",
    }

    for ri, row in enumerate(rows[1:], 2):
        category = row[0] if row else ""
        fgColor = category_colors.get(category, "FFFFFF")
        fill = PatternFill("solid", fgColor=fgColor)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            if headers[ci-1] in ('Planned_USD', 'Actual_USD') and val:
                try:
                    cell.value = float(val)
                    money_format(cell)
                except (ValueError, TypeError):
                    pass
            if 'TOTAL' in str(val).upper() or 'Milestone' in str(category):
                cell.font = Font(bold=True)

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(SCRIPT_DIR / "Budget.xlsx")
    print(f"  Saved Budget.xlsx ({len(rows)-1} rows)")


# ─── Timeline.xlsx ────────────────────────────────────────────────────────────

def create_timeline_xlsx():
    if not HAS_OPENPYXL:
        return
    print("Creating Timeline.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timeline"
    ws.freeze_panes = "A2"

    csv_path = SCRIPT_DIR / "Timeline.csv"
    if not csv_path.exists():
        print(f"  Warning: {csv_path} not found")
        return

    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))

    headers = rows[0]
    hdr = header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        apply_style(cell, hdr)

    phase_colors = {
        'Research':    "D9EAD3",
        'Design':      "CFE2F3",
        'Procurement': "FFF2CC",
        'Assembly':    "FCE4D6",
        'Integration': "EAD1DC",
        'RF_Test':     "D0E0E3",
        'Software':    "F4CCFF",
        'Field_Test':  "FFE599",
        'Documentation': "E8F5E9",
        'Milestone':   "FF9900",
    }

    status_colors = {
        'Complete':    "00B050",
        'Todo':        "808080",
        'In Progress': "FF9900",
    }

    for ri, row in enumerate(rows[1:], 2):
        phase = row[0] if row else ""
        status = row[7] if len(row) > 7 else ""
        fgColor = phase_colors.get(phase, "FFFFFF")
        fill = PatternFill("solid", fgColor=fgColor)

        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            if headers[ci-1] == 'Status' and val in status_colors:
                cell.font = Font(bold=True, color=status_colors[val])
            if headers[ci-1] == 'Duration_weeks' and val:
                try:
                    cell.value = int(val)
                except (ValueError, TypeError):
                    pass

    col_widths = {
        'Phase': 14, 'Task_ID': 8, 'Task': 50, 'Start_Date': 12,
        'End_Date': 12, 'Duration_weeks': 10, 'Dependencies': 12,
        'Status': 12, 'Notes': 50
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    wb.save(SCRIPT_DIR / "Timeline.xlsx")
    print(f"  Saved Timeline.xlsx ({len(rows)-1} tasks)")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("AERIS-10P Spreadsheet Generator")
    print("================================")
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        return

    create_bom_xlsx()
    create_budget_xlsx()
    create_timeline_xlsx()
    print("\nDone. Open .xlsx files in Microsoft Excel or LibreOffice Calc.")


if __name__ == '__main__':
    main()
